from openpyxl import load_workbook, Workbook
workbook = load_workbook(filename="score.xlsx")
workbook.sheetnames

sheet = workbook.active

sheet.title

# x = sheet["B5"].value
# print(x)
# print()

workbook = Workbook()  #ลบออก

for i in range(1,21):
    # print(i)
    print()
    sum = 0
    print(sum)
    print("g")
    Namei = []
    Namei.append(sheet["B"+str(i+4)].value)
    Namei.append(sheet["C"+str(i+4)].value)
    Namei.append(sheet["D"+str(i+4)].value)
    Namei.append(sheet["E"+str(i+4)].value)
    Namei.append(sheet["F"+str(i+4)].value)
    Namei.append(sheet["G"+str(i+4)].value)
    print("Name",i,"=",Namei,"1 =",Namei[0],"2 =",Namei[1],"3 =",Namei[2],"4 =",Namei[3],"5 =",Namei[4],"6 =",Namei[5])
    for k in range(6):
        sum = sum + Namei[k]
        #print(k)
    print("sum =",sum)
    sheet["I"+str(i+4)] = sum


# z = 0
# for k in range(1,21):
#     sumk = 0
#     print("Sum Name",i,"=",sumi)

workbook.save(filename="score.xlsx")