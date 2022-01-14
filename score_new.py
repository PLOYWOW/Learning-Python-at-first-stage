from openpyxl import load_workbook, Workbook
workbook = load_workbook(filename="score.xlsx")
workbook.sheetnames

sheet = workbook.active

sheet.title

Name1 = []
Name1.append(sheet["B5"].value)
Name1.append(sheet["C5"].value)
Name1.append(sheet["D5"].value)
Name1.append(sheet["E5"].value)
Name1.append(sheet["F5"].value)
Name1.append(sheet["G5"].value)

print(Name1)

sum = 0
for k in range(6):
    sum = sum + Name1[k]
print(sum)

sheet["I5"] = sum

workbook.save(filename="score.xlsx")